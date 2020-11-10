# the script uses a pandas dataframe to automate the filtering
# operation of a an excel file
# the filtering operation consist on a set of two filters in
# series: the first one defines a kind of item while the second one
# defines the number of those kinds whithin each of 4 different 
# groups. The table in output from the filtering is saved into an
# excel file properly formated.

import pandas as pd
import os
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Function found at:
        https://stackoverflow.com/questions/38074678/append-existing-excel-sheet-with-new-dataframe-using-python-pandas
    
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    
    from openpyxl import load_workbook

    import pandas as pd
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
#---------------------------------------------------------------------------------------------------------------------------    
#---------------------------------------------- script parameters ----------------------------------------------------------
path = "."
configuration_panel = 'prova.xlsx'
format_griglia = 'format.xlsx'
items = ['side','bottom','window']
groups = ['ic1','ic2','ie4']
#---------------------------------------------- script parameters ----------------------------------------------------------
#---------------------------------------------------------------------------------------------------------------------------    


# read the file
infile = path + '\\' + configuration_panel
dt = pd.read_excel(infile)

# prepare logical vector to extract informations
logic_rows1 = []
for item in items:
    logic_rows1.append([ dato.find(item) >=0 for dato in dt.iloc[:,2] ] ) 

logic_rows2 = []
for group in groups:
    logic_rows2.append( dt[group]>0 )
    
con1 = 0    
for logic_row1 in logic_rows1:
    con2 = 0
    for logic_row2 in logic_rows2:
        tabella = dt[logic_row1 & logic_row2] 
        griglia = tabella.iloc[:,[0,1,con2+3]]
        
        try:
            os.mkdir('.\\' + items[con1] )
        except:
            pass
        os.system('xcopy ".\\' + format_griglia + '" ' + '".\\' + items[con1] + '" /y')
        
        append_df_to_excel('.\\' + items[con1] + '\\' + format_griglia, griglia, startrow=3,header=None, index=False)
        
        # ------------------------------------------ write into the cells of excel file ---------------------------------------------------
        wb = load_workbook('.\\' + items[con1] + '\\' + format_griglia )
        ws = wb["Sheet1"]
        ws["A1"] = items[con1]
        ws["C3"] = groups[con2]
        wb.save('.\\' + items[con1] + '\\' + groups[con2] + '.xlsx')
        # ------------------------------------------ write into the cells of excel file ---------------------------------------------------
        
        os.remove('.\\' + items[con1] + '\\' + format_griglia )
        
        con2 += 1
    con1 += 1

